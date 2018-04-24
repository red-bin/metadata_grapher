#!/usr/bin/python3

import re
import psycopg2
from openpyxl import load_workbook

connstr = "dbname=metadata host=localhost user=metadata password=metadata"
conn = psycopg2.connect(connstr)
curs = conn.cursor()

def employee_email_addrs():
    sqlstr = """SELECT DISTINCT(addr) FROM
                  (SELECT sender_addr as addr 
                   FROM email_communications 
                   WHERE sender_addr like '%@houstontx.gov'
                   UNION
                   SELECT recip_addr as addr 
                   FROM email_communications
                   WHERE recip_addr like '%@houstontx.gov') as sq"""

    curs.execute(sqlstr)
    employee_addrs = [ r for r in curs.fetchall() ]

    return employee_addrs

def insert_employee(first_name=None, mid_name=None, last_name=None, 
                    title=None, department_id=None, hire_date=None, 
                    status=None, base_pay_rate=None, wage_type=None, 
                    email_address=None):
    sqlstr = """INSERT INTO employees
                (first_name, mid_name, last_name, title, department_id,
                 hire_date, status, base_pay_rate, wage_type, email_address)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

    vals = (first_name, mid_name, last_name, title, department_id, 
            hire_date, status, base_pay_rate, wage_type, email_address)

    curs.execute(sqlstr, vals)

    return

def insert_employees(sheet):                
    sheet_rows = sheet.rows

    header = [ h.value for h in sheet_rows.__next__() ]
    rows = (dict(zip(header, (c.value for c in r ))) for r in sheet_rows)

    for r in rows:
        insert_employee(first_name=r['First Name'].lower(),
          mid_name=r['Mid Name'].lower(),
          last_name=r['Last Name'].lower(),
          title=r['Title'],
          department_id=int(r['Department']),
          hire_date=r['Hire Date'],
          status=r['Status'],
          base_pay_rate=r['Base Pay Rate'],
          wage_type=r['Base Pay Wage Type'],
          email_address=None)

    conn.commit()

    return

def insert_departments(sheet):
    sheet_rows = sheet.rows

    header = [ h.value for h in sheet_rows.__next__() ]
    rows = (dict(zip(header, (c.value for c in r ))) for r in sheet_rows)


    sqlstr = """INSERT INTO departments (id, department_name) 
                VALUES (%s, %s)"""
    
    for r in rows:
        vals = (r['DEPT #'], r['DEPT'])
        curs.execute(sqlstr, vals)

    conn.commit()

    return
        
filepath = '/home/matt/git/metadata_grapher/houston.xlsx'

wb = load_workbook(filepath)

sheetnames = wb.sheetnames

employees_sheetname = 'Sheet1'
departments_sheetname = 'Dept Cross Table'

departments_sheet = wb[departments_sheetname]
insert_departments(departments_sheet)

employees_sheet = wb[employees_sheetname]
insert_employees(employees_sheet)


#addrs = employee_email_addrs()
#
#matches = []
#non_matches = []
#for addr in addrs:
#    addr = addr[0].lower()
#
#    email_first = addr.split('.')[0]
#    email_last = addr.split('@')[0].split('.')[-1]
#
#    addr_matches = []
#    for row in rows:
#        last_name = row[0].value.lower()
#        first_name = row[1].value.lower()
#
#        if email_first == first_name and email_last == last_name:
#            addr_matches.append((first_name, last_name, addr))
#
#    if addr_matches:
#        matches.append(addr_matches)
#
#    else:
###        non_matches.append(addr)
